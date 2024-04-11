from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill


def get_image_colors(image: Image) -> dict:
    """
    Retorna um dicionário com as cores presentes na imagem e suas coordenadas.

    Parâmetros:
    image (PIL.Image): A imagem a ser processada.

    Retorna:
    dict: Um dicionário onde as chaves são as cores encontradas na imagem e os valores são listas de coordenadas (x, y) onde essas cores ocorrem.
    """
    width, height = image.size
    pixels = image.load()
    colors_with_coordinates = {}

    # Verifica RGBA.
    if image.mode == 'RGBA':
        for y in range(height):
            for x in range(width):
                color = pixels[x, y]
                rgba_color = color + (pixels[x, y][3],)

                if rgba_color in colors_with_coordinates:
                    colors_with_coordinates[rgba_color]['coordenadas'].append({'x': x, 'y': y})
                else:
                    colors_with_coordinates[rgba_color] = {'coordenadas': [{'x': x, 'y': y}]}

    # Verifica RGB.
    else:
        for y in range(height):
            for x in range(width):
                color = pixels[x, y]

                if color in colors_with_coordinates:
                    colors_with_coordinates[color]['coordenadas'].append({'x': x, 'y': y})
                else:
                    colors_with_coordinates[color] = {'coordenadas': [{'x': x, 'y': y}]}

    return colors_with_coordinates


def write_image_in_excel(image: Image, output_path: str, coordinates: bool) -> None:
    """
    Escreve a imagem em um arquivo Excel, com a opção de incluir as coordenadas das cores.

    Parâmetros:
    image (PIL.Image): A imagem a ser escrita.
    output_path (str): O caminho do arquivo Excel de saída.
    coordinates (bool): Se True, as coordenadas das cores serão incluídas no arquivo Excel.

    Retorna:
    None
    """
    image = image.convert('RGBA')
    width, height = image.size

    wb = openpyxl.Workbook()
    ws = wb.active

    for y in range(height):
        for x in range(width):
            color = image.getpixel((x, y))

            # Substitui o preto pelo branco para melhor visualização.
            if color[:3] == (0, 0, 0):
                color = (255, 255, 255, 255)

            fill = PatternFill(start_color=f'{color[0]:02X}{color[1]:02X}{color[2]:02X}',
                               end_color=f'{color[0]:02X}{color[1]:02X}{color[2]:02X}', fill_type='solid')
            ws.cell(row=y + 1, column=x + 1).fill = fill

            if coordinates:
                ws.cell(row=y + 1, column=x + 1).value = f"{x},{y}"

    wb.save(output_path)
