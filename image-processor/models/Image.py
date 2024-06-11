from PIL import Image as PILImage
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np
from skimage.transform import resize as sk_resize, rescale as sk_rescale
from skimage import exposure, util


class Image:
    def __init__(self, image_path):
        self.image = PILImage.open(image_path)

    def get_image_colors(self):
        width, height = self.image.size
        pixels = self.image.load()
        colors_with_coordinates = {}

        if self.image.mode == 'RGBA':
            for y in range(height):
                for x in range(width):
                    color = pixels[x, y]
                    rgba_color = color + (pixels[x, y][3],)

                    if rgba_color in colors_with_coordinates:
                        colors_with_coordinates[rgba_color]['coordenadas'].append({'x': x, 'y': y})
                    else:
                        colors_with_coordinates[rgba_color] = {'coordenadas': [{'x': x, 'y': y}]}

        else:
            for y in range(height):
                for x in range(width):
                    color = pixels[x, y]

                    if color in colors_with_coordinates:
                        colors_with_coordinates[color]['coordenadas'].append({'x': x, 'y': y})
                    else:
                        colors_with_coordinates[color] = {'coordenadas': [{'x': x, 'y': y}]}

        return colors_with_coordinates

    def write_image_in_excel(self, output_path, coordinates=False):
        image = self.image.convert('RGBA')
        width, height = image.size

        wb = openpyxl.Workbook()
        ws = wb.active

        for y in range(height):
            for x in range(width):
                color = image.getpixel((x, y))

                if color[:3] == (0, 0, 0):
                    color = (255, 255, 255, 255)

                fill = PatternFill(start_color=f'{color[0]:02X}{color[1]:02X}{color[2]:02X}',
                                   end_color=f'{color[0]:02X}{color[1]:02X}{color[2]:02X}', fill_type='solid')
                ws.cell(row=y + 1, column=x + 1).fill = fill

                if coordinates:
                    ws.cell(row=y + 1, column=x + 1).value = f"{x},{y}"

        wb.save(output_path)

    def rescale_image(self, proportion):
        return sk_rescale(np.array(self.image), proportion)

    def resize_image(self, new_height, new_width):
        return np.array(sk_resize(np.array(self.image), (new_height, new_width), anti_aliasing=True))

    def flip_image_horizontally(self):
        return np.array(np.array(self.image)[:, ::-1, ...])

    def flip_image_vertically(self):
        return np.array(np.array(self.image)[::-1, :, ...])

    def mirror_image(self):
        flipped_image = self.flip_image_horizontally()
        return np.hstack([np.array(self.image), flipped_image])

    def move_image(self):
        pass

    def change_contrast(self, contrast_level):
        return exposure.adjust_gamma(np.array(self.image), gamma=contrast_level)

    def change_brightness(self, brightness_level):
        image_float = util.img_as_float(np.array(self.image))
        adjusted_image = image_float * brightness_level

        return exposure.adjust_gamma(adjusted_image, gamma=brightness_level)

def main():
    image = Image(image_path="../images/foto de perfil.jpg")