import cv2
from PIL import Image as PIL_Image, ImageEnhance, ImageOps
import openpyxl
from openpyxl.styles import PatternFill
import numpy as np
import io


class Image:
    def __init__(self, file):
        self.image = PIL_Image.open(file)

    def get_image_colors(self):
        width, height = self.image.size
        pixels = self.image.load()
        colors_with_coordinates = {}

        if self.image.mode == 'RGBA':
            for y in range(height):
                for x in range(width):
                    color = pixels[x, y]
                    rgba_color = color + (pixels[x, y][3],)

                    if rgba_color in colors_with_coordinates:
                        colors_with_coordinates[rgba_color]['coordenadas'].append({'x': x, 'y': y})
                    else:
                        colors_with_coordinates[rgba_color] = {'coordenadas': [{'x': x, 'y': y}]}

        else:
            for y in range(height):
                for x in range(width):
                    color = pixels[x, y]

                    if color in colors_with_coordinates:
                        colors_with_coordinates[color]['coordenadas'].append({'x': x, 'y': y})
                    else:
                        colors_with_coordinates[color] = {'coordenadas': [{'x': x, 'y': y}]}

        return colors_with_coordinates

    def write_image_in_excel(self, coordinates=False):
        image = self.image.convert('RGBA')
        width, height = image.size

        wb = openpyxl.Workbook()
        ws = wb.active

        for y in range(height):
            for x in range(width):
                color = image.getpixel((x, y))

                if color[:3] == (0, 0, 0):
                    color = (255, 255, 255, 255)

                fill = PatternFill(start_color=f'{color[0]:02X}{color[1]:02X}{color[2]:02X}',
                                   end_color=f'{color[0]:02X}{color[1]:02X}{color[2]:02X}', fill_type='solid')
                ws.cell(row=y + 1, column=x + 1).fill = fill

                if coordinates:
                    ws.cell(row=y + 1, column=x + 1).value = f"{x},{y}"

        # Cria um buffer em memória
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)  # Reposiciona o cursor no início do buffer

        return buffer

    def rescale_image(self, proportion):
        return self.image.resize(self.image.height * proportion, self.image.width * proportion)

    def resize_image(self, new_height, new_width):
        return self.image.resize((new_width, new_height), PIL_Image.Resampling.LANCZOS)

    def flip(self):
        return ImageOps.flip(self.image)

    def mirror_image(self):
        return ImageOps.mirror(self.image)

    def move_image(self):
        pass

    def rotate_image(self, value):
        return self.image.rotate(value, expand=True)

    def change_contrast(self, contrast_level):
        enhancer = ImageEnhance.Contrast(self.image)
        return enhancer.enhance(contrast_level)

    def change_brightness(self, brightness_level):
        enhancer = ImageEnhance.Brightness(self.image)
        return enhancer.enhance(brightness_level)

    def smooth_image(self, kernel_size=(5, 5), sigma=0):
        smoothed_image = cv2.GaussianBlur(np.array(self.image), kernel_size, sigma)
        return PIL_Image.fromarray(smoothed_image)

    def sharpen_image(self):
        kernel = np.array([[0, -1, 0],
                           [-1, 5, -1],
                           [0, -1, 0]])

        sharpened_image = cv2.filter2D(np.array(self.image), -1, kernel)
        sharpened_image_rgb = cv2.cvtColor(sharpened_image, cv2.COLOR_BGR2RGB)
        return PIL_Image.fromarray(sharpened_image_rgb)


def main():
    image = Image("../images/Reconhecimento_foto de perfil.jpg")
    image.sharpen_image()


if __name__ == "__main__":
    main()
